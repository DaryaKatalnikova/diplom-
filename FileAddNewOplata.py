from time import strptime
import main, onDataBase, Oplata
import openpyxl as op
from datetime import date, datetime

from PyQt6.QtCore import *
from PyQt6.QtWidgets import *
from pathlib import Path
from mysql.connector import Error

connection = onDataBase.create_connection("localhost", "root", "HarryPotterand3", "dbasit")

def execute_read_query(connection, query):
    cursor = connection.cursor()
    try:
        cursor.execute(query)
        result = cursor.fetchall()
        return result
    except Error as e:
        print(f"The error '{e}' occurred")

selectOC = "SELECT * FROM pay"
centers = execute_read_query(connection, selectOC)
if centers ==[]:
    id_pay = 1
else:
    id_pay = centers[len(centers)-1][0]
    id_pay += 1
'''
selectOC = "SELECT * FROM contracts"
centers = execute_read_query(connection, selectOC)
if centers ==[]:
    id_con = 1
else:
    id_con = centers[len(centers)-1][0]
    id_con += 1

'''
class AddWindowOplata(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("EducationPayTracker-Add-New-Oplata")
        # min_height = 480
        # min_weight = 480
        # self.setMinimumSize(min_weight, min_height)
        self.setGeometry(100, 100, 700, 250)
        self.setStyleSheet('''
                    QWidget{
                        background-color: qlineargradient(spread:pad, x1:1, y1:1, x2:0, y2:0, stop:0 rgba(217, 217, 217, 1), stop:0.427447 rgba(217, 217, 217, 1), stop:1 rgba(217, 217, 217, 1));
                        color: #D9D9D9;
                    }
                    QLabel{
                        font-family: Ubuntu Mono;
                        font-size: 18px;
                        color: #D9D9D9;
                    }
                    QMenuBar {
                        color: #000000;
                        border: 1px solid #424242;
                        border-radius: 7px;
                    }
                    QMenuBar::item {
                         spacing: 3px;
                         padding: 1px 4px;
                         background: transparent;
                         border-radius: 4px;
                     }
                    QLineEdit{
                        background-color: rgba(255, 255, 255, 30);
                        border: 1px solid #424242;
                        border-radius: 7px;
                        color: black;
                        font-weight: light;
                        font-size: 12pt;
                        padding-left: 10px;
                    }
                    QPushButton{
                        color: black;
                        background-color: rgba(255, 255, 255, 30);
                        border: 1px solid #424242;
                        border-radius: 7px;
                        width: 230px;
                        height: 30px;
                    }
                    QPushButton:hover{
                        background-color: rgba(255, 255, 255, 40);
                    }
                    QPushButton:pressed{
                        background-color: rgba(255, 255, 255, 70);
                    }
                    QTableWidget{
                        color: black;
                        border: 1px solid #424242;
                        border-bottom-right-radius: 7px;
                        border-bottom-left-radius: 7px;
                    }
                    QHeaderView::section{
                        Font-size:14px;
                        Font-family: "Microsoft YaHei";
                        Color: #FFFFFF;
                        Background:#60669B;
                        Border:none;
                        Text-align:left;
                        Min-height: 49px;
                        max-height:49px; 
                        Margin-left:0px;
                        Padding-left: 0px;
                    }
                ''')
        self.initUI()

    def initUI(self):
        # Основные кнопки
        self.choosetext = QLabel('Для того чтобы выбрать файл, нажмите на соответствующую кнопку')
        self.labelfile = QLabel('Файл:')
        self.filename_edit = QLineEdit(self)
        self.filename_edit.setPlaceholderText('Здесь отобразится пут файла')
        self.button_choose_oplata = QPushButton('Выбрать файл', self)
        self.button_upload_oplata_file = QPushButton('Загрузить файл', self)
        self.button_back_main_window_oplata = QPushButton('Вернуться на главное окно', self)

        # Сетка
        oplata_layout = QGridLayout()
        oplata_layout.addWidget(self.choosetext, 0, 0, 1, 3, alignment=Qt.AlignmentFlag.AlignCenter)
        oplata_layout.addWidget(self.labelfile, 2, 0, alignment=Qt.AlignmentFlag.AlignRight)
        oplata_layout.addWidget(self.filename_edit, 2, 1)
        oplata_layout.addWidget(self.button_choose_oplata, 3, 1)
        oplata_layout.addWidget(self.button_back_main_window_oplata, 5, 0)
        oplata_layout.addWidget(self.button_upload_oplata_file, 5, 2)
        oplata_layout.setRowStretch(1, 1)
        oplata_layout.setRowStretch(4, 1)
        self.setLayout(oplata_layout)

        self.button_back_main_window_oplata.clicked.connect(self.BackMainWindow)
        self.button_choose_oplata.clicked.connect(self.Choose_File)
        self.button_upload_oplata_file.clicked.connect(self.Upload_File)

    def BackMainWindow(self):
        self.o = main.WindowMain()
        self.o.show()
        self.close()

    def Choose_File(self):
        filename, _ = QFileDialog.getOpenFileName(
            self,
            "Select a File",
            "G:\Diplomchik2\ExcelAndProcject",
            "Excel (*.xlsx *.xls)"
        )
        if filename:
            path = Path(filename)
            self.filename_edit.setText(str(path))
        

    def Upload_File(self):
        try:
            connection = onDataBase.create_connection("localhost", "root", "HarryPotterand3", "dbasit")
            cursor = connection.cursor()
            filepath = " ".join(self.filename_edit.text().split()).title()
            wb = op.load_workbook(filepath, data_only=True)
            sheet = wb.active
            max_rows = sheet.max_row
            id_pay2 = id_pay
            for j in range(6, max_rows + 1):
                numberdoc = sheet.cell(row=j, column=1).value
                if numberdoc != None:
                    dateexcel = sheet.cell(row=j, column=2).value
                    namestudent = sheet.cell(row=j, column=3).value
                    namedogovor = sheet.cell(row=j, column=4).value
                    findelemdogovor = namedogovor.find('№')
                    simvolsdogovor = ""
                    for i in range(0, 8):
                        simvolsdogovor = simvolsdogovor + namedogovor[findelemdogovor]
                        findelemdogovor += 1
                    zapisal = simvolsdogovor
                    sum = sheet.cell(row=j, column=5).value
                    yearB = int(dateexcel.split('.')[2])
                    monthB = int(dateexcel.split('.')[1])
                    dayB = int(dateexcel.split('.')[0])

                    datefileaddnew = date(yearB, monthB, dayB)
                    connection = onDataBase.create_connection("localhost", "root", "HarryPotterand3", "dbasit")
                    select_sotrs = "SELECT * FROM student"
                    sotrs = execute_read_query(connection, select_sotrs)
                    lSotrs = list()
                    for s in sotrs:
                        lSotrs.append(str(s).split(',')[0][1:])
                    id_st_data_base = lSotrs[-1]
                    id_st = int(id_st_data_base) + 1
            
                    cursor = connection.cursor()
                    cursor.execute("SELECT secondname, namee, midlename, id_student FROM student")
                    result = list(cursor.fetchall())
                    for i in range(len(result)):
                        if namestudent in str(result[i][0])+' '+str(result[i][1])+' '+str(result[i][2]):
                            connection = onDataBase.create_connection("localhost", "root", "HarryPotterand3", "dbasit")
                            cursor = connection.cursor()
                            zapis = (id_pay2, sum, datefileaddnew, int(result[i][3]))
                            newaddoplata = "INSERT INTO pay (id_pay, summa, pay_date, id_student) values (%s, %s, %s, %s)"
                            cursor.execute(newaddoplata, zapis)
                            connection.commit()
                            id_pay2 += 1
                            connection.close()
        except:
            print("Путь не выбран")
        self.close()
        self.n = Oplata.Oplata()
        self.n.showMaximized()